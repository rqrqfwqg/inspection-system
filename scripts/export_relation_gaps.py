"""导出「关联缺口清单」：把 build_relations.py 解析不到的引用列出来，供现场补充。
- 复用与 build_relations.py 完全一致的 norm / resolve 逻辑
- 输出 scripts/relation_gaps.xlsx：
     sheet1「缺口清单」：来源表/设备编号/指向机房(原始)/归一化键/机房主表是否存在/现场补充机房编号(待填)
     sheet2「机房主表有效编号」：所有可作为目标的机房编号 + 名称/楼栋/楼层（现场比对用）
- 控制台打印缺口统计（按来源表、按归一化前缀聚类），便于判断缺失范围
用法：python export_relation_gaps.py
"""
import sys, os, re
from collections import Counter
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + "/../backend")
from database import SessionLocal, DataTable, Record, Subsystem, FieldDef

db = SessionLocal()

def t(code):
    return db.query(DataTable).filter(DataTable.code == code).first()

def norm(s):
    if not s:
        return None
    s = s.strip().upper()
    s = re.sub(r'^([A-Z]+?)(\d)', r'\1-\2', s)   # GW2F -> GW-2F
    s = re.sub(r'-\d+$', '', s)                   # 去尾部 -N
    return s

# ---- 机房主表索引（与 build_relations.py 同）----
rt = t("room_master")
rooms = db.query(Record).filter(Record.table_id == rt.id).all()
room_codes = [r.device_code for r in rooms if r.device_code]
room_norm = {}
for rc in room_codes:
    room_norm.setdefault(norm(rc), []).append(rc)

def resolve(room_no):
    if not room_no:
        return None
    rn = room_no.strip()
    if rn in room_codes:
        return rn
    nk = norm(rn)
    cands = room_norm.get(nk, [])
    if not cands:
        return None
    if re.search(r'-\d+$', rn):
        suffix = rn.rsplit('-', 1)[-1]
        for c in cands:
            if c.endswith('-' + suffix):
                return c
    return sorted(cands)[0]

# 机房主表字段标签
def fld_label(tb, key):
    f = db.query(FieldDef).filter(FieldDef.table_id == tb.id, FieldDef.key == key).first()
    return f.label if f else key

def fld_val(rec, key):
    return (rec.data or {}).get(key)

# 关联源定义（与 build_relations.py 一致）
LINKS = [
    ("power_cabinets", "room_no", "power", "所在机房"),
    ("ba_integrated_ac", "room", "hvac", "所在机房"),
]

gap_rows = []
for table_code, field, sub_code, rel_type in LINKS:
    tb = t(table_code)
    if not tb:
        continue
    sub = db.query(Subsystem).filter(Subsystem.code == sub_code).first()
    sub_name = sub.name if sub else sub_code
    fld = fld_label(tb, field)
    recs = db.query(Record).filter(Record.table_id == tb.id).all()
    for r in recs:
        child = r.device_code
        ref = fld_val(r, field)
        if not child or not ref:
            continue
        parent = resolve(ref)
        if parent:
            continue  # 能匹配，不列入缺口
        gap_rows.append({
            "来源子系统": sub_name,
            "来源资料表": tb.name,
            "来源字段": fld,
            "设备编号": child,
            "指向机房(原始)": ref,
            "归一化键": norm(ref) or "",
            "机房主表是否存在": "否",
            "现场补充机房编号": "",        # 待现场填写
            "处理方式": "补机房主表后重跑 / 或直接建关联",
        })

# ---- 缺口统计 ----
by_src = Counter((g["来源资料表"]) for g in gap_rows)
by_prefix = Counter(g["归一化键"] for g in gap_rows)
print(f"=== 关联缺口统计：共 {len(gap_rows)} 条未匹配 ===")
for src, n in by_src.most_common():
    print(f"  {src}: {n} 条")
print("\n--- 缺口按归一化键聚类（前 25，揭示缺失机房类型）---")
for pfx, n in by_prefix.most_common(25):
    print(f"  {pfx or '(空)'}: {n}")

# ---- 写 Excel ----
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "缺口清单"
headers = list(gap_rows[0].keys()) if gap_rows else ["（无缺口）"]
ws.append(headers)
for g in gap_rows:
    ws.append([g[h] for h in headers])
# 自动列宽
from openpyxl.utils import get_column_letter
for i, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(h) + 4))

# sheet2：机房主表有效编号（现场比对）
ws2 = wb.create_sheet("机房主表有效编号")
rm_building = fld_label(rt, "building")
rm_floor = fld_label(rt, "floor")
rm_name = fld_label(rt, "room_name")
ws2.append(["机房编号", rm_name, rm_building, rm_floor])
for r in rooms:
    ws2.append([r.device_code, fld_val(r, "room_name"), fld_val(r, "building"), fld_val(r, "floor")])
for i, h in enumerate(["机房编号", rm_name, rm_building, rm_floor], 1):
    ws2.column_dimensions[get_column_letter(i)].width = 20

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "relation_gaps.xlsx")
wb.save(out)
print(f"\n已导出：{out}")
db.close()
